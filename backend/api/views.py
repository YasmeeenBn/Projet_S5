from django.shortcuts import render
from bs4 import BeautifulSoup
from urllib.request import Request, urlopen
from .models import Article, Thematic
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import ArticleSerializer
from django.http import HttpResponse
from django.http import JsonResponse
import os
import csv
import json
from django.core import serializers
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.db.models import Q
from rest_framework.views import APIView
from rest_framework import filters
from rest_framework import generics


#export data from CSV
def export_csv(request):
    response = HttpResponse(content_type='text/csv')
    writer = csv.writer(response)
    writer.writerow(['title', 'date', 'imageUrl', 'link'])
    #map on Articles to bring data and write it in the csv file
    for article in Article.objects.all().values_list('title', 'date', 'imageUrl', 'link'):
        writer.writerow(article)
    #put the data in the csv file
    response['Content-Disposition'] = 'attachment; filename="articles.csv"'
    return response

#export data form json
def export_json(request):
    articles = Article.objects.all()
    articles = serializers.serialize('json', articles) 
    return HttpResponse(articles,content_type="application/json")

#export data form json
def export_thematics_hespress_json(request):
    thematics = Thematic.objects.filter(website = 'https://en.hespress.com/')
    thematics = serializers.serialize('json', thematics) 
    return HttpResponse(thematics,content_type="application/json")

def export_thematics_ny_json(request):
    thematics = Thematic.objects.filter(website = 'https://www.nytimes.com/')
    thematics = serializers.serialize('json', thematics) 
    return HttpResponse(thematics,content_type="application/json")


#export data form xml
def export_xml(request):
    articles = Article.objects.all()
    articles = serializers.serialize('xml', articles) 
    return HttpResponse(articles,content_type="application/xml")

#export data form xlsx
def export_xlsx(request):
    article = Article.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-Articles.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Articles'
    # Define the titles for columns
    columns = [
        'ID',
        'Title',
        'Link',
        'ImageUrl',
        'Date',
    ]
    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # Iterate through all movies
    for article in article:
        row_num += 1
        # Define the data for each cell in the row 
        row = [
            article.id,
            article.title,
            article.link,
            article.imageUrl,
            article.date,
        ]
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response


#api 
@api_view(['GET'])
def articleList(request):

    articles = Article.objects.all()
    serilizer = ArticleSerializer(articles, many=True) #json
    # print(artic)
    return Response(serilizer.data)

class ListVariable(APIView):
    def post(self, request, format=None):
        serializer = VariableSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class Articlefilter(generics.ListAPIView):

    queryset = Article.objects.all()
    serializer_class = ArticleSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['^title','^thematic']
    # '^' Starts-with search.
    # '=' Exact matches.
    # '@' Full-text search. (Currently only supported Django's PostgreSQL backend.)
    # '$' Regex search.
     

@api_view(['POST'])
def search(request):

    query = request.data.get('query', '')

    if query:
        article = Article.objects.filter(Q(title__icontains=query))
        # article = Article.objects.filter(Q(title__icontains=query) | Q(description__icontains=query))
        serializer = ArticleSerializer(article, many=True)
        return Response(serializer.data)
    else:
        return Response({"Articles": []})


